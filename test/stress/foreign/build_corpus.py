"""
Tier B: the corpus the Node writers cannot produce.

Everything in test/stress/generators/_write.ts is written by DuckDB or
apache-arrow, because those are what the extension already depends on and what
CI can run without Python. That covers a great deal, and it has one blind spot
that has already cost real time: a corpus written by the same library that
reads it agrees with the reader by construction.

`malformedFiles.test.ts` passed for months while every polars-written Feather
file failed, because every Arrow fixture in it came from DuckDB. DuckDB writes
plain `Utf8`; polars writes `Utf8View`, which `read_arrow` rejects outright.
Nothing in the suite could produce the failing shape, so nothing did.

This script writes the same logical table through the writers the viewer
actually meets in the wild -- polars, pyarrow, openpyxl, pandas and DuckDB's
own Python bindings -- and records what it produced in manifest.json. The TS
side reads that manifest; cases whose files are absent report SKIPPED with the
command to build them, and never fail. So `npm test` stays green on a machine
with no Python, and CI needs no Python step, while a developer who runs this
gets several hundred extra cases over files nobody in this repo could
otherwise write.

Run it with the project's env:

    conda run -n myproject python test/stress/foreign/build_corpus.py

Output goes to test/stress/_work/foreign/, which is gitignored. The manifest is
committed, so a missing file is distinguishable from a case nobody wrote.
"""

from __future__ import annotations

import json
import shutil
import sys
from datetime import date, datetime
from pathlib import Path

HERE = Path(__file__).resolve().parent
OUT = HERE.parent / "_work" / "foreign"
MANIFEST = HERE / "manifest.json"

# One logical table, written every way. Kept small on purpose: these files
# exist to vary the WRITER, not the data, and a difference that only shows up
# at 100,000 rows is a different question from the one being asked here.
ROWS = [
    (1, "alpha", 1.5, date(2020, 1, 31), True),
    (2, "beta", 2.5, date(2020, 2, 29), False),
    (3, None, 3.5, date(2020, 3, 31), True),
    (4, "İstanbul ısı", -4.5, date(2020, 4, 30), False),
    (5, "has, comma", 0.0, date(2020, 5, 31), None),
]
COLUMNS = ["id", "label", "amount", "day", "flag"]

# What every one of these files must read back as, in the JSON space the
# viewer's grid uses. The TS side compares against this, so it lives here
# beside the writers rather than being restated per case.
EXPECTED = {
    "columns": COLUMNS,
    "rows": [
        [1, "alpha", 1.5, "2020-01-31", True],
        [2, "beta", 2.5, "2020-02-29", False],
        [3, None, 3.5, "2020-03-31", True],
        [4, "İstanbul ısı", -4.5, "2020-04-30", False],
        [5, "has, comma", 0.0, "2020-05-31", None],
    ],
}

entries: list[dict] = []
skipped: list[str] = []


def record(path: Path, writer: str, note: str, *, expected: dict | None = EXPECTED) -> None:
    entries.append(
        {
            "file": path.name,
            "writer": writer,
            "note": note,
            "expected": expected,
        }
    )


def _frame():
    import polars as pl

    return pl.DataFrame(
        {
            "id": [r[0] for r in ROWS],
            "label": [r[1] for r in ROWS],
            "amount": [r[2] for r in ROWS],
            "day": [r[3] for r in ROWS],
            "flag": [r[4] for r in ROWS],
        }
    )


def build_polars() -> None:
    import polars as pl

    df = _frame()

    # THE case this whole tier exists for. polars' default compat level writes
    # string columns as Utf8View, which read_arrow rejects as "Unrecognized
    # Field type with value 24" -- so the container conversion is not enough on
    # its own and the columns have to be downcast too.
    df.write_ipc(OUT / "polars-default.feather", compat_level=pl.CompatLevel.newest())
    record(OUT / "polars-default.feather", "polars", "default compat level: Utf8View strings")

    df.write_ipc(OUT / "polars-oldest.feather", compat_level=pl.CompatLevel.oldest())
    record(OUT / "polars-oldest.feather", "polars", "oldest compat level: plain Utf8 strings")

    # The IPC STREAM encoding, which is what read_arrow reads natively -- and
    # what macro_project and ETL.py are told to write, so it is the shape the
    # user's own pipelines produce.
    df.write_ipc_stream(OUT / "polars-stream.arrows", compat_level=pl.CompatLevel.oldest())
    record(OUT / "polars-stream.arrows", "polars", "IPC stream, the encoding the R/Python pipelines emit")

    df.write_parquet(OUT / "polars.parquet")
    record(OUT / "polars.parquet", "polars", "parquet written by polars rather than DuckDB")

    df.write_csv(OUT / "polars.csv")
    record(OUT / "polars.csv", "polars", "csv written by polars")

    # Compressed Feather. apache-arrow JS ships no IPC codecs, so these CANNOT
    # be converted and must be refused with a message that says so and points
    # at the way out -- not surfaced as "codec not found".
    for codec in ("lz4", "zstd"):
        path = OUT / f"polars-{codec}.feather"
        df.write_ipc(path, compression=codec, compat_level=pl.CompatLevel.oldest())
        record(path, "polars", f"{codec}-compressed Feather, which must be refused clearly", expected=None)


def build_pyarrow() -> None:
    import pyarrow as pa
    import pyarrow.feather as feather
    import pyarrow.parquet as pq

    table = pa.table(
        {
            "id": pa.array([r[0] for r in ROWS], pa.int32()),
            "label": pa.array([r[1] for r in ROWS], pa.string()),
            "amount": pa.array([r[2] for r in ROWS], pa.float64()),
            "day": pa.array([r[3] for r in ROWS], pa.date32()),
            "flag": pa.array([r[4] for r in ROWS], pa.bool_()),
        }
    )

    feather.write_feather(table, OUT / "pyarrow-v2.feather", version=2, compression="uncompressed")
    record(OUT / "pyarrow-v2.feather", "pyarrow", "Feather V2, uncompressed")

    # Feather V1 is a wholly different container -- not Arrow IPC at all -- and
    # still turns up in files written years ago.
    try:
        feather.write_feather(table, OUT / "pyarrow-v1.feather", version=1)
        record(
            OUT / "pyarrow-v1.feather",
            "pyarrow",
            "Feather V1, a different container entirely; refused or read, never half-read",
            expected=None,
        )
    except Exception as exc:  # pragma: no cover - depends on the pyarrow build
        skipped.append(f"pyarrow v1: {exc}")

    # Dictionary-encoded strings: a different Arrow layout for the same values,
    # and one nothing in the Node corpus produces.
    dict_table = table.set_column(
        table.schema.get_field_index("label"),
        "label",
        table.column("label").dictionary_encode(),
    )
    feather.write_feather(dict_table, OUT / "pyarrow-dictionary.feather", version=2, compression="uncompressed")
    record(OUT / "pyarrow-dictionary.feather", "pyarrow", "dictionary-encoded string column")

    # Many small batches, so the batch-at-a-time conversion has boundaries to
    # get wrong.
    batched = pa.Table.from_batches(table.to_batches(max_chunksize=1))
    feather.write_feather(batched, OUT / "pyarrow-batched.feather", version=2, compression="uncompressed")
    record(OUT / "pyarrow-batched.feather", "pyarrow", "one row per record batch")

    pq.write_table(table, OUT / "pyarrow.parquet")
    record(OUT / "pyarrow.parquet", "pyarrow", "parquet written by pyarrow")

    # Parquet compressions DuckDB should all handle; here to prove it rather
    # than assume it.
    for codec in ("snappy", "gzip", "zstd", "brotli"):
        path = OUT / f"pyarrow-{codec}.parquet"
        try:
            pq.write_table(table, path, compression=codec)
            record(path, "pyarrow", f"{codec}-compressed parquet")
        except Exception as exc:  # pragma: no cover - codec availability varies
            skipped.append(f"parquet {codec}: {exc}")


def build_openpyxl() -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(COLUMNS)
    for row in ROWS:
        ws.append(list(row))
    # A second sheet, so "the edit touched only one part" is a real assertion.
    other = wb.create_sheet("notes")
    other.append(["do not touch me"])
    wb.save(OUT / "openpyxl.xlsx")
    record(
        OUT / "openpyxl.xlsx",
        "openpyxl",
        "a workbook laid out by openpyxl: shared strings, a real styles part, a dimension element",
        # openpyxl writes the date column as a formatted serial, which read_xlsx
        # surfaces as a timestamp rather than a date -- a legitimate difference
        # between writers, and the reason this one declares no expected table.
        expected=None,
    )

    # The shape that broke header detection on the real workbook: notes under
    # the table, so counting back from the row count lands inside the data.
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "data"
    ws2.append(COLUMNS)
    for row in ROWS:
        ws2.append(list(row))
    ws2.append([])
    ws2.append(["Source: internal"])
    ws2.append(["Revised 2024"])
    wb2.save(OUT / "openpyxl-trailing-notes.xlsx")
    record(
        OUT / "openpyxl-trailing-notes.xlsx",
        "openpyxl",
        "notes below the table, written by a real library rather than by hand",
        expected=None,
    )

    # A merged title banner above the header -- what most human-made workbooks
    # look like, and the shape that collapses a sheet to one column.
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = "data"
    ws3.append(["Quarterly figures"])
    ws3.merge_cells("A1:E1")
    ws3.append(COLUMNS)
    for row in ROWS:
        ws3.append(list(row))
    wb3.save(OUT / "openpyxl-merged-title.xlsx")
    record(
        OUT / "openpyxl-merged-title.xlsx",
        "openpyxl",
        "a merged title banner above the header",
        expected=None,
    )


def build_pandas() -> None:
    import pandas as pd

    df = pd.DataFrame(ROWS, columns=COLUMNS)
    df.to_csv(OUT / "pandas.csv", index=False)
    record(OUT / "pandas.csv", "pandas", "csv written by pandas, including its own NA spelling", expected=None)

    try:
        # version=118 is the first Stata format with UTF-8 string storage. The
        # default (117) is latin-1, which cannot hold `İ` at all -- so without
        # this the one kind that has NO Tier A source, because DuckDB's dta
        # extension reads but does not write, would be skipped over a character
        # that appears in the user's real column names.
        # Stata has no date-object type; it wants a datetime64 column with an
        # explicit conversion. Done on a copy so the CSV above keeps the plain
        # dates -- the point of Tier B is that each writer's own idea of the
        # data reaches the viewer, not that they are normalised to agree.
        # Stata has neither a date-object type nor a boolean one, so both need
        # an explicit conversion. Done on a copy: the point of Tier B is that
        # each writer's own idea of the data reaches the viewer, not that the
        # writers are normalised into agreeing with each other.
        stata = df.copy()
        stata["day"] = pd.to_datetime(stata["day"])
        stata["flag"] = stata["flag"].map({True: 1, False: 0})
        stata.to_stata(OUT / "pandas.dta", write_index=False, version=118)
        record(
            OUT / "pandas.dta",
            "pandas",
            "Stata .dta — DuckDB's dta extension reads but cannot write, so there is no Tier A source at all",
            expected=None,
        )
    except Exception as exc:
        skipped.append(f"pandas dta: {exc}")


def build_duckdb_python() -> None:
    import duckdb

    path = OUT / "duckdb-python.duckdb"
    if path.exists():
        path.unlink()
    con = duckdb.connect(str(path))
    con.execute(
        """
        create table data (id INTEGER, label VARCHAR, amount DOUBLE, day DATE, flag BOOLEAN)
        """
    )
    con.executemany("insert into data values (?, ?, ?, ?, ?)", [list(r) for r in ROWS])
    con.close()
    record(path, "duckdb-python", "a database written by the Python bindings, read by the Node ones")


def main() -> int:
    if OUT.exists():
        shutil.rmtree(OUT)
    OUT.mkdir(parents=True, exist_ok=True)

    builders = [
        ("polars", build_polars),
        ("pyarrow", build_pyarrow),
        ("openpyxl", build_openpyxl),
        ("pandas", build_pandas),
        ("duckdb", build_duckdb_python),
    ]

    for name, fn in builders:
        try:
            fn()
        except ImportError as exc:
            skipped.append(f"{name}: not installed ({exc})")
        except Exception as exc:
            # A writer that breaks must not cost the rest of the corpus.
            skipped.append(f"{name}: {type(exc).__name__}: {exc}")

    MANIFEST.write_text(
        json.dumps(
            {
                "generated": datetime.now().isoformat(timespec="seconds"),
                "command": "conda run -n myproject python test/stress/foreign/build_corpus.py",
                "files": sorted(entries, key=lambda e: e["file"]),
            },
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )

    print(f"wrote {len(entries)} file(s) to {OUT}")
    for entry in sorted(entries, key=lambda e: e["file"]):
        size = (OUT / entry["file"]).stat().st_size
        print(f"  {entry['file']:36} {size:>9,} B  {entry['writer']}")
    for line in skipped:
        print(f"  skipped: {line}")
    print(f"manifest: {MANIFEST}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
